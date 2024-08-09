# -*-coding: utf-8 -*-
# @Time    : 2024/5/19 15:43
# @Author  : 张银昌
# @File    : 美国物流距离分割算法.py
# @Software: PyCharm
#物流解决方案
import requests
from datetime import datetime
from googlemaps import Client
from  wlgjsc.工具信息 import Tools
from wlgjsc.长距离分割算法 import Get_cjlfg
from openpyxl import Workbook
from wlgjsc.数据库操作 import  Mysql_gssbdb
class Get_wlinxfo(object):
     def __init__(self):
         self.wb = Workbook()
         self.ws = self.wb.active
         self.cjjlfg_data=Get_cjlfg()
         self.tools=Tools()
         self.API_KEY = 'AIzaSyBt9hH7SC3rJu2eAchw_A-zt4Qrg2ic0Gs'
         self.client = Client(key=self.API_KEY)
         self.url = 'https://maps.googleapis.com/maps/api/directions/json'
         self.proxy = {'http': '127.0.0.1:7897', 'https': '127.0.0.1:7897'}
         self.lanage = "en-hk" 'zh-CN'
         self.parm = {
             "origin": "37.6273584, -122.1077314",
             "destination": "33.7646983, -84.5315202",
             "language": self.lanage,
             "key": self.API_KEY
         }
         pass

     def get_ldjl(self, origin, destination,hpbh,cysbh,time_string,zsc):
         self.ws['A1']="起点经纬度"
         self.ws['B1'] = "起点地址"
         self.ws['C1'] = "行驶距离"
         self.ws['D1'] = "行驶速度"
         self.ws['E1'] = "行驶用时"
         self.ws['F1'] = "抵达时间"
         self.ws['G1'] = "抵达地址经纬度"
         self.ws['H1'] = "抵达地址"


         self.parm['origin'] = origin
         self.parm['destination'] = destination
         if self.tools.getis_sc()=="N":
            response = requests.get(self.url, params=self.parm, proxies=self.proxy)
         else:
            response = requests.get(self.url, params=self.parm)
         print(response.text)
         dis = {}
         if response.status_code == 200:
             data = response.json()
             if data['status'] == 'OK':
                 route = data['routes'][0]
                 dis['起始位置'] = route['legs'][0]['start_address']
                 dis['结束位置'] = route['legs'][0]['end_address']
                 dis['总距离'] = route['legs'][0]['distance']['value']
                 dis['API总时长'] = route['legs'][0]['duration']['value']
                 dis['总时长'] =zsc
                 dis['每秒钟速度(米/秒)']= round(  dis['总距离']/dis['总时长'],2)
                 dis['60分钟的距离'] = round(dis['每秒钟速度(米/秒)']*60*60,2)
                 dis['60分钟的距离'] = 60*60
                 # now = datetime.now()
                 time1 = datetime.strptime(time_string, '%Y-%m-%d %H:%M:%S')
                 print(f'time1{time1}')
                 dis['开始时间'] = datetime.strptime(time_string, '%Y-%m-%d %H:%M:%S') #now.strftime("%Y-%m-%d %H:%M:%S")
                 print(dis)
                 lis = route['legs'][0]['steps']
                 totoljl = 0
                 for index, li in enumerate(lis):
                     # url = f"https://maps.googleapis.com/maps/api/distancematrix/json?units=metric&language=zh-CN&key={self.API_KEY}&origins={li['start_location']['lat']},{li['start_location']['lng']}&destinations={li['end_location']['lat']},{li['end_location']['lng']}"
                     # res = requests.get(url, proxies=self.proxy)
                     if 'distance' in li and li['distance'] and li['distance']['text'] != '':
                         #行驶的总距离 总距离在超过
                         print(f"60分钟的距离{dis['60分钟的距离']} --{li['distance']['value']}")
                         if dis['60分钟的距离'] > li['distance']['value']:


                             disqsdz={}

                             disqsdz['latitude']=li['start_location']['lat']
                             disqsdz['longitude'] = li['start_location']['lng']
                             qswz=self.getdz_tojwd(disqsdz) #

                             disjsdz = {}
                             disjsdz['latitude'] = li['end_location']['lat']
                             disjsdz['longitude'] = li['end_location']['lng']
                             jswz=self.getdz_tojwd(disjsdz)
                             if index == 0:
                                 self.start_time = str(dis['开始时间'])
                             # print(f'开始时间{type(self.start_time)}')
                             self.start_time=datetime.strptime(self.start_time, '%Y-%m-%d %H:%M:%S')
                             distance = li['distance']['value']
                             qsjwd=f"{disqsdz['latitude']},{disqsdz['longitude']}"
                             jsjwd = f"{disjsdz['latitude']},{disjsdz['longitude']}"
                             self.start_time = self.tools.calculate_arrival_time(self.start_time, distance, dis['每秒钟速度(米/秒)'])
                             self.start_time = str(self.start_time)[:19]
                             dis['开始时间']=self.start_time
                             ret = f"起点经纬度:{disqsdz} --起点地址{qswz} 行驶{li['distance']['value']}米--行驶速度{dis['每秒钟速度(米/秒)']}--用时{li['duration']['value']}秒--抵达时间{self.start_time}--抵达地址经纬度 {disjsdz} --抵达地址{jswz}"
                             totoljl+=li['distance']['value']
                             if totoljl>=30000 :
                                 totoljl=0
                             else:
                                 continue;
                             print(ret)
                             lis_data=[qsjwd,qswz,li['distance']['value'],dis['每秒钟速度(米/秒)'],li['duration']['value'],self.start_time,jsjwd,jswz]
                             print(f"lis_data--{lis_data}")


                             self.ws=self.tools.setexl(self.ws,lis_data,'0',hpbh,cysbh) #>30
                             print(f"{dis['60分钟的距离']}--{li['distance']['value']}--{totoljl}")
                             print(ret)

                         if     dis['60分钟的距离'] <= li['distance']['value']:
                                disqsdz['latitude'] = li['start_location']['lat']
                                disqsdz['longitude'] = li['start_location']['lng']
                                qswz = self.getdz_tojwd(disqsdz)  #
                                disjsdz['latitude'] = li['end_location']['lat']
                                disjsdz['longitude'] = li['end_location']['lng']
                                jswz = self.getdz_tojwd(disjsdz)  #
                                dis['开始时间'] = self.start_time
                                dis['origin']=qswz
                                dis['destination'] = jswz
                                decret=self.cjjlfg_data.get_cjldada(dis,self.ws,'1',hpbh,cysbh)#<30
                                self.start_time=decret['start_time']
                                self.ws = decret['ws']
             self.wb.save("轨迹信息-21062.xlsx")


     def getdz_tojwd(self,disorg):
         reverse_geocode_result = self.client.reverse_geocode((disorg['latitude'], disorg['longitude']), language='zh-CN')
         address = reverse_geocode_result[0].get('formatted_address')
         return address
if __name__ == '__main__':
    wlinfo=Get_wlinxfo()

    Mysql_db = Mysql_gssbdb()
    Mysql_db.del_Logisticstrajectory('001','002')
    # origin = '温斯洛 亚利桑那州邮政编码: 86047'  # 例如，Los Angeles
    # destination = '3PCJ+XM, Mesita, NM'  # 例如，Vancouver

    origin = '21062 Forbes Ave Hayward CA 94545'  # 例如，Los Angeles
    destination = '1115 Wesel Blvd Hagerstown MD 21740'  # 例如，Vancouver
    time_string='2024-07-15 00:00:00'
    zsc=172800.0
    origin='兰州大学'
    destination='天津大学'


    wlinfo.get_ldjl(origin,destination,'01','01',time_string,zsc)
